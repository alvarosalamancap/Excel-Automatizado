try:
    import openpyxl
    from openpyxl.styles import Font, Border, Side
except ImportError:
    print("La biblioteca openpyxl no está instalada. Se procederá con la instalación.")
    import subprocess
    subprocess.run(["pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, Border, Side

def crear_tabla_excel():
    # Preguntar al usuario por el número de filas y columnas
    num_filas = int(input("Ingrese el número de filas: "))
    num_columnas = int(input("Ingrese el número de columnas: "))

    # Preguntar si desea los nombres de las columnas en negrita
    formato_negrita = input("¿Quiere los nombres de las columnas en negrita? (Sí:S / No:N): ").lower() == 's'

    # Crear un nuevo libro de trabajo y seleccionar la hoja activa
    libro = openpyxl.Workbook()
    hoja = libro.active

    # Preguntar al usuario por los nombres de las columnas y escribirlos en la primera fila
    for col_num in range(1, num_columnas + 1):
        nombre_columna = input(f"Ingrese el nombre de la columna {col_num}: ")

        # Aplicar formato negrita si es seleccionado
        if formato_negrita:
            hoja.cell(row=1, column=col_num, value=nombre_columna).font = Font(bold=True)
        else:
            hoja.cell(row=1, column=col_num, value=nombre_columna)

    # Preguntar al usuario por los nombres de las filas y escribirlos en la primera columna
    for fila_num in range(2, num_filas + 2):
        nombre_fila = input(f"Ingrese el nombre de la fila {fila_num - 1}: ")
        hoja.cell(row=fila_num, column=1, value=nombre_fila)

    # Preguntar si desea agregar bordes
    agregar_bordes = input("¿Quiere agregar bordes a la tabla? (Sí:S / No:N): ").lower() == 's'

    # Agregar bordes si es seleccionado
    if agregar_bordes:
        for fila in hoja.iter_rows(min_row=1, max_row=num_filas + 1, min_col=1, max_col=num_columnas):
            for celda in fila:
                celda.border = Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )

    # Guardar el archivo Excel
    libro.save('tabla_excel_generada.xlsx')

    print("Tabla Excel generada con éxito.")

if __name__ == "__main__":
    crear_tabla_excel()
